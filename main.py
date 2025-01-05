from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from gtts import gTTS
from moviepy import ImageClip, AudioFileClip, concatenate_videoclips
from pathlib import Path
import os

class VideoCreator:
    def __init__(self, excel_file: str, font_path: str, output_video: str):
        self.excel_file = excel_file
        self.font_path = font_path
        self.output_video = output_video
        
        # Настройки видео
        self.frame_size = (1920, 1080)
        self.font_size = 60
        self.text_color = "black"
        self.bg_color = "white"
        self.fade_duration = 1
        self.fps = 24
        
        # Создаем временные директории
        self.frames_dir = Path("frames")
        self.audio_dir = Path("audio")
        self.frames_dir.mkdir(exist_ok=True)
        self.audio_dir.mkdir(exist_ok=True)
        
        # Инициализируем шрифт
        self.font = ImageFont.truetype(font_path, self.font_size)
        
    def load_texts(self):
        """Загружаем тексты из Excel файла"""
        workbook = load_workbook(self.excel_file)
        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        return [{"de": row[0], "ru": row[1]} for row in rows if row[0] and row[1]]
    
    def create_frame(self, de_text: str, ru_text: str, idx: int) -> str:
        """Создаем кадр с текстом"""
        img = Image.new("RGB", self.frame_size, color=self.bg_color)
        draw = ImageDraw.Draw(img)
        
        # Немецкий текст
        de_bbox = draw.textbbox((0, 0), de_text, font=self.font)
        de_width = de_bbox[2] - de_bbox[0]
        de_height = de_bbox[3] - de_bbox[1]
        de_position = ((self.frame_size[0] - de_width) // 2,
                      self.frame_size[1] // 2 - de_height - 20)
        draw.text(de_position, de_text, font=self.font, fill=self.text_color)
        
        # Русский текст
        ru_bbox = draw.textbbox((0, 0), ru_text, font=self.font)
        ru_width = ru_bbox[2] - ru_bbox[0]
        ru_height = ru_bbox[3] - ru_bbox[1]
        ru_position = ((self.frame_size[0] - ru_width) // 2,
                      self.frame_size[1] // 2 + 20)
        draw.text(ru_position, ru_text, font=self.font, fill=self.text_color)
        
        frame_path = self.frames_dir / f"frame_{idx}.png"
        img.save(frame_path)
        return str(frame_path)
    
    def create_video(self):
        """Создаем видео из кадров и аудио"""
        texts = self.load_texts()
        clips = []
        
        for idx, text_pair in enumerate(texts):
            de_text, ru_text = text_pair["de"], text_pair["ru"]
            
            # Создаем аудио
            tts = gTTS(text=f"{de_text}. {ru_text}", lang="de")
            audio_path = self.audio_dir / f"audio_{idx}.mp3"
            tts.save(str(audio_path))
            audio_clip = AudioFileClip(str(audio_path))
            
            # Создаем кадр
            frame_path = self.create_frame(de_text, ru_text, idx)
            
            # Создаем видеоклип
            video_clip = (ImageClip(frame_path)
                         .with_duration(audio_clip.duration + 2)  # +2 секунды для fade эффектов
                         .fadein(self.fade_duration)
                         .fadeout(self.fade_duration)
                         .set_audio(audio_clip))
            
            clips.append(video_clip)
        
        # Собираем финальное видео
        final_video = concatenate_videoclips(clips, method="compose")
        final_video.write_videofile(self.output_video, fps=self.fps, codec="libx264")
        
        # Закрываем клипы
        final_video.close()
        for clip in clips:
            clip.close()
    
    def cleanup(self):
        """Удаляем временные файлы"""
        for frame in self.frames_dir.glob("*"):
            frame.unlink()
        for audio in self.audio_dir.glob("*"):
            audio.unlink()
        self.frames_dir.rmdir()
        self.audio_dir.rmdir()

def main():
    # Параметры
    excel_file = "phrases.xlsx"
    font_path = "C:/Windows/Fonts/arial.ttf"
    output_video = "output.mp4"
    
    # Создаем и запускаем видео
    creator = VideoCreator(excel_file, font_path, output_video)
    try:
        creator.create_video()
        print(f"Видео успешно сохранено в {output_video}")
    except Exception as e:
        print(f"Произошла ошибка: {e}")
    finally:
        creator.cleanup()

if __name__ == "__main__":
    main()